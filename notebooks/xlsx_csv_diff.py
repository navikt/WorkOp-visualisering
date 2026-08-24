# ---
# jupyter:
#   jupytext:
#     formats: py:percent
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.16.0
#   kernelspec:
#     display_name: Python 3
#     language: python
#     name: python3
# ---

# %% [markdown]
# # Xlsx vs. CSV — datasammenligning for WO 1–46
#
# WO 1–46 ble opprinnelig registrert i Excel-arket
# `Resultat og måling 1-46 Workop.xlsx`. Dataene ble manuelt overført
# til to Forms-CSV-filer. Denne notebooken sjekker om det er avvik.

# %%
import re
import sys
import warnings
from pathlib import Path

import pandas as pd

warnings.filterwarnings("ignore")
sys.path.insert(0, "..")

DATA_DIR = Path("..") / "data"
XLSX_PATH = DATA_DIR / "Resultat og måling 1-46 Workop.xlsx"
SEMANTIKK_GRENSE = 46

# %% [markdown]
# ## 1. Les xlsx-data (gammel kilde)

# %%
import mimetypes

_real_mime_read = mimetypes.MimeTypes.read

def _safe_mime_read(self, filename, strict=True):
    try:
        return _real_mime_read(self, filename, strict)
    except (PermissionError, OSError):
        pass

mimetypes.MimeTypes.read = _safe_mime_read  # type: ignore[method-assign]

import openpyxl  # noqa: E402

WORKOP_PATTERN = re.compile(r"^WorkOp\s+(\d+)\s*$")


def _safe_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    m = re.search(r"\d+", str(value))
    return int(m.group()) if m else None


def _get_active_sheet_names(wb):
    names = wb.sheetnames
    try:
        start = names.index("Start") + 1
        stop = names.index("Slutt")
    except ValueError:
        start, stop = 0, len(names)
    return [n for n in names[start:stop] if WORKOP_PATTERN.match(n)]


def xlsx_extract_all(filepath: Path) -> pd.DataFrame:
    """Leser nøkkeltall fra hvert WorkOp-ark i xlsx-fila."""
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    sheet_names = _get_active_sheet_names(wb)

    rows = []
    for name in sheet_names:
        ws = wb[name]
        nr_match = WORKOP_PATTERN.match(name)
        workop_nr = int(nr_match.group(1)) if nr_match else None
        rows.append({
            "workop_nr": workop_nr,
            "oppmotte_forberedende": _safe_int(ws["E4"].value),
            "oppmotte": _safe_int(ws["E5"].value),
            "arbeidsgivere": _safe_int(ws["E9"].value),
            "innkalt_intervju": _safe_int(ws["E13"].value),
            "fatt_jobb": _safe_int(ws["E14"].value),
            "fatt_jobb_tiltak": _safe_int(ws["E15"].value),
            "fatt_jobb_nedsatt": _safe_int(ws["F14"].value),
            "fatt_jobb_veiledning": _safe_int(ws["G14"].value),
            "fatt_jobb_gode": _safe_int(ws["H14"].value),
        })

    return pd.DataFrame(rows).sort_values("workop_nr").reset_index(drop=True)


# %%
df_xlsx = xlsx_extract_all(XLSX_PATH)
print(f"Xlsx: {len(df_xlsx)} WorkOp-er lest")
df_xlsx.head()

# %% [markdown]
# ## 2. Les CSV-data (nåværende kilde)

# %%
from src.workop.extract import extract_all

df_csv_full, csv_warns = extract_all()
df_csv = df_csv_full[df_csv_full["workop_nr"] <= SEMANTIKK_GRENSE].copy()
# Dedupliser: test-rader i Forms2 skaper duplikater (feks WO 46).
# Behold raden med høyest fatt_jobb per workop_nr.
df_csv = (df_csv.sort_values("fatt_jobb", ascending=False, na_position="last")
                .drop_duplicates(subset="workop_nr", keep="first")
                .sort_values("workop_nr")
                .reset_index(drop=True))
print(f"CSV: {len(df_csv)} WorkOp-er (WO ≤ {SEMANTIKK_GRENSE}, deduplisert)")
if csv_warns:
    print(f"CSV-advarsler: {csv_warns}")

# %% [markdown]
# ## 3. Sammenlign nøkkeltall

# %%
COMPARE_FIELDS = [
    "oppmotte_forberedende",
    "oppmotte",
    "arbeidsgivere",
    "innkalt_intervju",
    "fatt_jobb",
    "fatt_jobb_tiltak",
    "fatt_jobb_nedsatt",
    "fatt_jobb_veiledning",
    "fatt_jobb_gode",
]

# For CSV bruker vi fatt_jobb direkte (allerede beregnet med semantikk-skift for WO ≤ 46)
# og fatt_jobb_tiltak = ansatt_med_tiltak (kompatibilitetskolonne)

merged = df_xlsx.merge(df_csv[["workop_nr"] + COMPARE_FIELDS],
                        on="workop_nr", how="outer",
                        suffixes=("_xlsx", "_csv"))

diffs = []
for _, row in merged.iterrows():
    wo = int(row["workop_nr"])
    for field in COMPARE_FIELDS:
        xlsx_val = row.get(f"{field}_xlsx")
        csv_val = row.get(f"{field}_csv")
        # Normaliser NaN/None
        if pd.isna(xlsx_val):
            xlsx_val = None
        else:
            xlsx_val = int(xlsx_val)
        if pd.isna(csv_val):
            csv_val = None
        else:
            csv_val = int(csv_val)

        if xlsx_val != csv_val:
            # None vs 0 er ofte bare at xlsx manglet feltet — flagg separat
            er_none_vs_0 = (xlsx_val is None and csv_val == 0) or (xlsx_val == 0 and csv_val is None)
            diffs.append({
                "workop_nr": wo,
                "felt": field,
                "xlsx": xlsx_val,
                "csv": csv_val,
                "diff": (csv_val or 0) - (xlsx_val or 0),
                "none_vs_0": er_none_vs_0,
            })

df_diffs = pd.DataFrame(diffs)

# Filtrer bort None-vs-0-avvik som standard
VIS_NONE_VS_0 = False  # Sett til True for å inkludere None↔0-avvik
df_reelle = df_diffs[~df_diffs["none_vs_0"]] if not df_diffs.empty else df_diffs

if df_reelle.empty:
    print("✅ Ingen reelle avvik i nøkkeltall mellom xlsx og CSV for WO 1–46")
else:
    print(f"⚠️  {len(df_reelle)} reelle avvik funnet i nøkkeltall:\n")
    print(df_reelle.drop(columns="none_vs_0").to_string(index=False))

if not df_diffs.empty:
    n_none_0 = df_diffs["none_vs_0"].sum()
    if n_none_0 > 0:
        print(f"\n({n_none_0} None↔0-avvik skjult — sett VIS_NONE_VS_0 = True for å vise)")

# %% [markdown]
# ### Oppsummering per felt

# %%
if not df_reelle.empty:
    print("Avvik per felt (uten None↔0):")
    print(df_reelle.groupby("felt").size().to_string())
    print(f"\nAntall WO-er med minst ett reelt avvik: {df_reelle['workop_nr'].nunique()}")
else:
    print("Ingen avvik å oppsummere.")

# %% [markdown]
# ## 4. Viktigste avvik: fått jobb

# %%
if not df_reelle.empty:
    jobb_felt = ["fatt_jobb", "fatt_jobb_tiltak", "fatt_jobb_nedsatt",
                 "fatt_jobb_veiledning", "fatt_jobb_gode"]
    jobb_diffs = df_reelle[df_reelle["felt"].isin(jobb_felt)]
    if jobb_diffs.empty:
        print("✅ Ingen reelle avvik i jobb-relaterte felt")
    else:
        print(f"⚠️  {len(jobb_diffs)} reelle avvik i jobb-felt:\n")
        print(jobb_diffs.drop(columns="none_vs_0").to_string(index=False))
else:
    print("Ingen avvik.")

# %% [markdown]
# ## 5. Totaloppsummering

# %%
n_total = len(df_reelle) if not df_reelle.empty else 0
n_wo = df_reelle["workop_nr"].nunique() if not df_reelle.empty else 0
n_none_0 = int(df_diffs["none_vs_0"].sum()) if not df_diffs.empty else 0

print("=" * 60)
print("OPPSUMMERING — xlsx vs. CSV nøkkeltall for WO 1–46")
print("=" * 60)
print(f"Reelle avvik:    {n_total} i {n_wo} WO-er")
print(f"None↔0 (skjult): {n_none_0}")
if n_total == 0:
    print("\n✅ Xlsx og CSV er konsistente for WO 1–46")
else:
    print("\n⚠️  Det finnes avvik — se detaljene over")
