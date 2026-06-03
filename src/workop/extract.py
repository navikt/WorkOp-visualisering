"""
Henter og normaliserer data fra WorkOp Excel-fila.

Bruk:
    df, missing = extract_all("data/Resultat og måling 1-32 Workop.xlsx")
"""

from __future__ import annotations

import mimetypes
import re
from datetime import datetime
from pathlib import Path

# --- macOS / Python 3.14 workaround -------------------------------------------
# openpyxl importerer MimeTypes() ved import, som på macOS prøver å lese
# /etc/apache2/mime.types og feiler med PermissionError.
# Patch MimeTypes.read *før* openpyxl importeres.
_real_mime_read = mimetypes.MimeTypes.read


def _safe_mime_read(self, filename, strict=True):  # type: ignore[override]
    try:
        return _real_mime_read(self, filename, strict)
    except (PermissionError, OSError):
        pass


mimetypes.MimeTypes.read = _safe_mime_read  # type: ignore[method-assign]
# ------------------------------------------------------------------------------

import openpyxl  # noqa: E402  (must come after patch)
import dateparser  # noqa: E402
import pandas as pd  # noqa: E402
from openpyxl.workbook.workbook import Workbook  # noqa: E402
from openpyxl.worksheet.worksheet import Worksheet  # noqa: E402

YEAR_SEPARATOR = "2026"
WORKOP_PATTERN = re.compile(r"^WorkOp\s+(\d+)\s*$")

# B1-verdier som er generiske plassholdere uten sted/dato-info
_GENERIC_TITTEL = re.compile(r"^WorkOp\s*(tid/sted:?)?\s*$", re.IGNORECASE)

# Arbeidsgiver-plassholdere ("Arbeidsgiver x", "Arbeidsgiver 1" o.l.)
PLACEHOLDER_AG = re.compile(r"^[Aa]rbeidsgiver[\s\dxX!]*$")
# Tall-only = feil i rad 31 (WorkOp 25/37-stil)
_GARBAGE_AG = re.compile(r"^\d+$")
# Bransje-verdier som er tall eller header-tekst
_GARBAGE_BRANSJE = re.compile(r"^\d+$|^Yrkeskategorier:?$", re.IGNORECASE)

# Normalisering fritekst-bransje → standardisert kategori (nøkler er lowercase).
# Oppdater når nye bransjetyper dukker opp.
BRANSJE_NORM: dict[str, str] = {
    # Restaurant og servering
    "restaurant": "Restaurant og servering",
    "servering": "Restaurant og servering",
    "servering/sjåfør": "Restaurant og servering",
    "servering/kokk": "Restaurant og servering",
    "servitør": "Restaurant og servering",
    "kafe": "Restaurant og servering",
    "rest./kantine/kafe": "Restaurant og servering",
    "restaurantmedarbeider og kokk/kjøkkenassistent": "Restaurant og servering",
    "medarbeider innen mat og servering": "Restaurant og servering",
    "overnattings- og serveringsvirksomhet": "Restaurant og servering",
    "kiosk/lager/servering": "Restaurant og servering",
    # Varehandel og butikk
    "varehandel": "Varehandel og butikk",
    "butikkmedarbeider": "Varehandel og butikk",
    "butikk": "Varehandel og butikk",
    "butikk/dagligvarehandel": "Varehandel og butikk",
    "butikkmedarbeider byggvare": "Varehandel og butikk",
    "butikkmedarbeider dagligvare": "Varehandel og butikk",
    "dagligvarehandel": "Varehandel og butikk",
    "detaljhandel": "Varehandel og butikk",
    "matbutikk": "Varehandel og butikk",
    "møbelbutikk": "Varehandel og butikk",
    "varehus": "Varehandel og butikk",
    "varehus (åpner i juni)": "Varehandel og butikk",
    "varehandel (detaljhandel med bredt vareutvalg)": "Varehandel og butikk",
    "varehandel/sport": "Varehandel og butikk",
    "engros": "Varehandel og butikk",
    "salg": "Varehandel og butikk",
    "salg/service": "Varehandel og butikk",
    "kioskmedarbeider/lager": "Varehandel og butikk",
    "bensinstasjon": "Varehandel og butikk",
    "07-eleven": "Varehandel og butikk",
    # Sikkerhet og vakt
    "sikkerhet": "Sikkerhet og vakt",
    "vekter": "Sikkerhet og vakt",
    "alarmoperatør": "Sikkerhet og vakt",
    "forretningsmessig tjenesteyting": "Sikkerhet og vakt",
    "forretningsmessig tjenesteyting (vakttjenester ikke nevnt annet sted)": "Sikkerhet og vakt",
    "saferoad": "Sikkerhet og vakt",
    "veisikkerhet": "Sikkerhet og vakt",
    "sikkerhet på vei": "Sikkerhet og vakt",
    "sikkerhet vei": "Sikkerhet og vakt",
    "trafikkdirigenter": "Sikkerhet og vakt",
    "vakthald/": "Sikkerhet og vakt",
    # Industri og produksjon
    "industri": "Industri og produksjon",
    "hjelpearbeider industri innen iso-faget": "Industri og produksjon",
    "industri (overflatebehandling av metaller)": "Industri og produksjon",
    "industri - fiskefor": "Industri og produksjon",
    "produksjon": "Industri og produksjon",
    "mat produksjon": "Industri og produksjon",
    "næringsmiddel": "Industri og produksjon",
    "sjømat": "Industri og produksjon",
    "lager/produksjon": "Industri og produksjon",
    "flycatering/lager/produksjon": "Industri og produksjon",
    "nor tekstil": "Industri og produksjon",
    # Renhold
    "renhold": "Renhold",
    "renhold/husøkonom": "Renhold",
    "forretningsmessig tjenesteyting (rengjøring av bygninger)": "Renhold",
    "elis": "Renhold",
    "housmen": "Renhold",
    # Lager og logistikk
    "lager": "Lager og logistikk",
    "lager/logistikk": "Lager og logistikk",
    "lagermedarbeider byggvare": "Lager og logistikk",
    "terminalarbeider": "Lager og logistikk",
    # Transport
    "transport": "Transport",
    "transport og service": "Transport",
    # Helse og omsorg
    "helsefagarbeider": "Helse og omsorg",
    "helse": "Helse og omsorg",
    "helse/oppvekst": "Helse og omsorg",
    "bpa": "Helse og omsorg",
    "hjemmetjenester/bpa": "Helse og omsorg",
    "oppvekst": "Helse og omsorg",
    # Barnehage og utdanning
    "barnehage": "Barnehage og utdanning",
    "undervisning/formidling": "Barnehage og utdanning",
    "tjønnstuggu barnehage": "Barnehage og utdanning",
    # Bygg og anlegg
    "bygg og anlegg": "Bygg og anlegg",
    "byggebransje": "Bygg og anlegg",
    "byggvarehus": "Bygg og anlegg",
    "tømrer / snekker": "Bygg og anlegg",
    "verksted": "Bygg og anlegg",
    # Hotell og reiseliv
    "hotell": "Hotell og reiseliv",
    "hotell/reiseliv": "Hotell og reiseliv",
    "hotellbransjen": "Hotell og reiseliv",
    "friluft": "Hotell og reiseliv",
    # Bil og kjøretøy
    "bil": "Bil og kjøretøy",
    "bilglass": "Bil og kjøretøy",
    "bilindustri": "Bil og kjøretøy",
    "(truck)mekaniker": "Bil og kjøretøy",
    "bilpleie": "Bil og kjøretøy",
    # Annet (beholder, men sorteres sist i plott)
    "it support": "Annet",
    "kontor og administrasjon": "Annet",
    "renovasjon": "Annet",
    "gjenvinning": "Annet",
    "bemanningsbyrå": "Annet",
    "jordbruk/gartneri": "Annet",
    "landbrukstjenester": "Annet",
    "telefonsalg": "Annet",
    "salgs- og markedskonsulent": "Annet",
    "låsesmed": "Annet",
    "foliering": "Annet",
}

NORSK_MAANEDER: dict[str, int] = {
    "jan": 1, "januar": 1,
    "feb": 2, "februar": 2,
    "mar": 3, "mars": 3,
    "apr": 4, "april": 4,
    "mai": 5,
    "jun": 6, "juni": 6,
    "jul": 7, "juli": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "okt": 10, "oktober": 10,
    "nov": 11, "november": 11,
    "des": 12, "desember": 12,
}

DATEPARSER_SETTINGS = {
    "RETURN_AS_TIMEZONE_AWARE": False,
}


def _load_workbook(filepath: str | Path) -> Workbook:
    return openpyxl.load_workbook(str(filepath), data_only=True)


def _get_active_sheet_names(wb: Workbook) -> list[str]:
    """Returnerer WorkOp-ark mellom 'Start' og 'Slutt' (eksklusiv)."""
    names = wb.sheetnames
    try:
        start = names.index("Start") + 1
        stop = names.index("Slutt")
    except ValueError:
        start, stop = 0, len(names)
    return [n for n in names[start:stop] if WORKOP_PATTERN.match(n)]


def _year_separator_index(wb: Workbook) -> int:
    """Indeks til '2026'-skillearket i sheetnames. Returnerer sys.maxsize om det mangler."""
    try:
        return wb.sheetnames.index(YEAR_SEPARATOR)
    except ValueError:
        return len(wb.sheetnames)


def _infer_fallback_year(sheet_name: str, wb: Workbook) -> int:
    sep_idx = _year_separator_index(wb)
    sheet_idx = wb.sheetnames.index(sheet_name)
    return 2026 if sheet_idx > sep_idx else 2025


def _les_tittel(ws: Worksheet) -> str | None:
    """
    Leser tittelcellen robust. Noen ark har tittelen i feil kolonne:
      - Standard: B1
      - Kolonneskift til høyre (WorkOp 11, 31): C1
      - Kolonneskift til venstre (WorkOp 19): A1
      - Kolonneskift med D (WorkOp 21): D1
      - Generisk B1 + sted/dato i D1 (WorkOp 14, 20, 36):
        B1 = 'WorkOp tid/sted', D1 = 'Tønsberg 05.02.2026'
        → kombinerer til 'WorkOp tid/sted Tønsberg 05.02.2026'

    Rekkefølge: B1 (hvis ikke generisk), deretter D1, C1, A1.
    Hvis B1 er generisk og D1 har innhold, kombineres de for full dato-parsbarhet.
    """
    b1_raw = ws.cell(1, 2).value
    b1 = str(b1_raw).strip() if b1_raw else ""

    if b1 and not _GENERIC_TITTEL.match(b1):
        return b1

    for col in [4, 3, 1]:  # D, C, A
        val = ws.cell(1, col).value
        if val and str(val).strip():
            extra = str(val).strip()
            return f"{b1} {extra}".strip() if b1 else extra

    return b1 or None


def _parse_date(raw: str | None, fallback_year: int) -> datetime | None:
    """
    Parser dato fra fritekst-strenger som 'WorkOp tid/sted 30.nov 2023'.

    Strategi:
      1. Finn eksplisitt 4-sifret år (2020–2029) eller kortår (23–26) i strengen.
         Mangler år → bruk fallback_year.
      2. Finn dag + måned via regex med norske månedsnavn.
      3. Fallback: rent numerisk format dd.mm.yy.
    """
    if not raw:
        return None

    s = str(raw).lower()

    # Finn eksplisitt 4-sifret år (2020–2029).
    # Kortår (25, 26 osv.) håndteres kun i det numeriske dd.mm.yy-mønsteret nedenfor,
    # for å unngå at dagnummer som "25" i "25.mars" feiltolkes som år.
    m_year4 = re.search(r"\b(202[0-9])\b", s)
    year = int(m_year4.group(1)) if m_year4 else fallback_year

    # Finn dag + måned: "30.nov", "7mars", "15. mai", "13. november"
    for mname in sorted(NORSK_MAANEDER, key=len, reverse=True):
        mnum = NORSK_MAANEDER[mname]
        # dag foran månedsnavn: "30.nov", "7mars", "25.feb", "13. november"
        m = re.search(rf"\b(\d{{1,2}})\.?\s*{mname}", s)
        if m:
            try:
                return datetime(year, mnum, int(m.group(1)))
            except ValueError:
                pass
        # månedsnavn foran dag (sjeldent): "november 13"
        m = re.search(rf"{mname}\.?\s*(\d{{1,2}})\b", s)
        if m:
            try:
                return datetime(year, mnum, int(m.group(1)))
            except ValueError:
                pass

    # Rent numerisk: dd.mm.yy[yy]
    m = re.search(r"\b(\d{1,2})\.(\d{1,2})\.(\d{2,4})\b", s)
    if m:
        day, month, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 100:
            yr = 2000 + yr
        try:
            return datetime(yr, month, day)
        except ValueError:
            pass

    return None


def _safe_int(value: object) -> int | None:
    """Konverterer til int. Håndterer None, float og strenger som '24 innkalt'."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    m = re.search(r"\d+", str(value))
    return int(m.group()) if m else None


def _safe_float(value: object) -> float | None:
    """Konverterer til float. Håndterer norsk tusenskille ('25.000' og '25 000')."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[\s.]", "", str(value)).replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _extract_workop(ws: Worksheet, sheet_name: str, wb: Workbook) -> dict:
    nr_match = WORKOP_PATTERN.match(sheet_name)
    workop_nr = int(nr_match.group(1)) if nr_match else None

    raw_title = _les_tittel(ws)
    title_kort = re.sub(r"^WorkOp\s+|tid/sted\s*", "", raw_title or "").strip()
    fallback_year = _infer_fallback_year(sheet_name, wb)
    dato = _parse_date(raw_title, fallback_year)

    fatt_jobb = _safe_int(ws["E14"].value)

    return {
        "workop_nr": workop_nr,
        "raw_title": raw_title,
        "title_kort": title_kort,
        "dato": dato,
        "fallback_year": fallback_year,
        "oppmotte_forberedende": _safe_int(ws["E4"].value),
        "oppmotte": _safe_int(ws["E5"].value),
        "arbeidsgivere": _safe_int(ws["E9"].value),
        "innkalt_intervju": _safe_int(ws["E13"].value),
        "fatt_jobb": fatt_jobb,
        "fatt_jobb_tiltak": _safe_int(ws["E15"].value),
        "fatt_jobb_nedsatt": _safe_int(ws["F14"].value),
        "fatt_jobb_veiledning": _safe_int(ws["G14"].value),
        "fatt_jobb_gode": _safe_int(ws["H14"].value),
        "dagsverk": _safe_float(ws["E7"].value),
        "kostnader": _safe_float(ws["E8"].value),
        "har_data": fatt_jobb is not None,
    }


def _validate_sum(df: pd.DataFrame, wb: Workbook) -> None:
    """Sammenligner vår sum(fatt_jobb) mot 'Sammenstilling av data' C2."""
    try:
        ws_sum = wb["Sammenstilling av data"]
        excel_total = _safe_int(ws_sum["C2"].value)
        our_total = df["fatt_jobb"].sum()
        if excel_total is not None and our_total != excel_total:
            print(
                f"⚠️  Sanity check: sum(fatt_jobb) = {our_total}, "
                f"Excel-sammenstilling viser {excel_total} "
                f"(diff = {our_total - excel_total})"
            )
        else:
            print(f"✅ Sanity check OK: sum(fatt_jobb) = {our_total}")
    except Exception:
        pass


def extract_all(filepath: str | Path) -> tuple[pd.DataFrame, list[str]]:
    """
    Leser alle aktive WorkOp-ark fra Excel-fila.

    Returnerer:
        df       — DataFrame med én rad per WorkOp
        missing  — liste over arknavn uten E14-data
    """
    wb = _load_workbook(filepath)
    sheet_names = _get_active_sheet_names(wb)

    rows = []
    for name in sheet_names:
        ws = wb[name]
        rows.append(_extract_workop(ws, name, wb))

    df = pd.DataFrame(rows)
    df = df.sort_values("workop_nr").reset_index(drop=True)

    missing_rows = df.loc[~df["har_data"], ["workop_nr", "raw_title"]].copy()
    base = missing_rows["raw_title"].fillna(
        missing_rows["workop_nr"].astype(int).astype(str).radd("WorkOp ")
    )
    missing = (
        missing_rows["workop_nr"].astype(int).astype(str).radd("Ark ")
        + " - "
        + base.astype(str)
    ).tolist()

    df_with_data = df[df["har_data"]].copy()
    _validate_sum(df_with_data, wb)

    return df, missing


# ---------------------------------------------------------------------------
# Arbeidsgiverdata — rad 31–37, én kolonne per arbeidsgiver
# ---------------------------------------------------------------------------

def _normaliser_bransje(raw: str | None) -> str | None:
    """
    Mapper råbransje til standardisert kategori via BRANSJE_NORM.
    Returnerer 'Annet' for kjente men ikke-kartlagte verdier,
    None for ugyldige/tomme verdier (tall, header-tekst, o.l.).
    """
    if not raw:
        return None
    raw = raw.strip()
    if _GARBAGE_BRANSJE.match(raw):
        return None
    # Lange kommalister er sannsynligvis feildata (flere firmaer på én rad)
    if len(raw) > 80 and raw.count(",") >= 2:
        return None
    return BRANSJE_NORM.get(raw.lower(), "Annet")


def _storrelse_kategori(antall_ansatte: int | None) -> str | None:
    """EU SME-standardkategorier basert på antall ansatte."""
    if antall_ansatte is None:
        return None
    if antall_ansatte < 10:
        return "Mikro"
    if antall_ansatte < 50:
        return "Liten"
    if antall_ansatte < 250:
        return "Medium"
    return "Stor"


def _extract_ag_kolonner(ws: Worksheet, workop_nr: int) -> list[dict]:
    """
    Leser arbeidsgiverdata fra rad 31–37, kolonne F (6) og utover.

    Stopper ved første tomme celle i rad 31. Hopper over kolonner der
    rad 31 inneholder tall-only (WorkOp 25/37-stil feildata).
    """
    result = []
    for col in range(6, 32):  # F → AF (maks ~26 arbeidsgivere)
        navn_raw = ws.cell(31, col).value
        if not navn_raw or not str(navn_raw).strip():
            break  # Stopp ved første tomme celle

        navn = str(navn_raw).strip()
        if _GARBAGE_AG.match(navn):
            continue  # Tall-only — skip uten å bryte løkken

        bransje_raw = ws.cell(32, col).value
        bransje_str = str(bransje_raw).strip() if bransje_raw else None

        antall_ansatte = _safe_int(ws.cell(33, col).value)

        result.append({
            "workop_nr": workop_nr,
            "ag_idx": len(result) + 1,
            "navn": navn,
            "er_anonym": bool(PLACEHOLDER_AG.match(navn)),
            "bransje_raw": bransje_str,
            "bransje": _normaliser_bransje(bransje_str),
            "antall_ansatte": antall_ansatte,
            "storrelse": _storrelse_kategori(antall_ansatte),
            "rekrutteringsbehov": _safe_int(ws.cell(34, col).value),
            "speedintervjuer": _safe_int(ws.cell(35, col).value),
            "aktuell": _safe_int(ws.cell(36, col).value),
            "ansatt": _safe_int(ws.cell(37, col).value),
        })
    return result


def extract_arbeidsgivere(filepath: str | Path) -> tuple[pd.DataFrame, list[str]]:
    """
    Henter arbeidsgiverdata fra alle aktive WorkOp-ark (rad 31–37).

    Returnerer:
        df_ag      — DataFrame med én rad per arbeidsgiver per WorkOp (lang format)
        ag_missing — liste over arknavn uten arbeidsgiverdata i rad 31
    """
    wb = _load_workbook(filepath)
    sheet_names = _get_active_sheet_names(wb)

    rows: list[dict] = []
    missing: list[str] = []

    for name in sheet_names:
        ws = wb[name]
        nr_match = WORKOP_PATTERN.match(name)
        if not nr_match:
            continue
        workop_nr = int(nr_match.group(1))

        ag_rows = _extract_ag_kolonner(ws, workop_nr)
        if ag_rows:
            rows.extend(ag_rows)
        else:
            missing.append(name)

    if rows:
        df_ag = pd.DataFrame(rows).sort_values(["workop_nr", "ag_idx"]).reset_index(drop=True)
    else:
        df_ag = pd.DataFrame(columns=[
            "workop_nr", "ag_idx", "navn", "er_anonym", "bransje_raw",
            "bransje", "antall_ansatte", "storrelse",
            "rekrutteringsbehov", "speedintervjuer", "aktuell", "ansatt",
        ])

    return df_ag, missing
